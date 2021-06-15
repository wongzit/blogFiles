# CSIgen v1.1, 2021-04-02
# Computational supporting information generator powered by Python 3.9
# Written by Zhe Wang, Hiroshima university
# Catch me with wongzit@yahoo.co.jp
# Personal webpage: https://www.wangzhe95.net

import openpyxl
from openpyxl.styles import Border, Font, Alignment, Side

# Program information section
print("*******************************************************************************")
print("*                                                                             *")
print("*                                  C S I g e n                                *")
print("*                                                                             *")
print("*     =================== Version 1.1 for Source Code ===================     *")
# Choose platform (for packaging)
#print("*     ====================== Version 1.1 for macOS ======================     *")
#print("*     ====================== Version 1.1 for Linux ======================     *")
#print("*     ================ Version 1.1 for Microsoft Windows ================     *")
print("*                           Last update: 2021-04-02                           *")
print("*                                                                             *")
print("*       Computational Supporting Information (CSI) generator, developed       *")
print("*      by Zhe Wang. Online document is available from GitHub.                 *")
print("*                              (https://github.com/wongzit/CSIgen)            *")
print("*                                                                             *")
print("*                             -- Catch me with --                             *")
print("*                         E-mail  wongzit@yahoo.co.jp                         *")
print("*                       Homepage  https://www.wangzhe95.net                   *")
print("*                                                                             *")
print("*******************************************************************************")
print("\nPRESS Ctrl+c to exit the program.\n")

def elementNo (eleNumber):
	element = 'H'
	periodTable = ['Bq', 'H', 'He', 'Li', 'Be', 'B', 'C', 'N', 'O', 'F', 'Ne', 'Na', 'Mg', 'Al', 'Si', 'P', 'S', 'Cl', 'Ar', \
					'K', 'Ca', 'Sc', 'Ti', 'V', 'Cr', 'Mn', 'Fe', 'Co', 'Ni', 'Cu', 'Zn', 'Ga', 'Ge', 'As', 'Se', 'Br', 'Kr', \
					'Rb', 'Sr', 'Y', 'Zr', 'Nb', 'Mo', 'Tc', 'Ru', 'Rh', 'Pd', 'Ag', 'Cd', 'In', 'Sn', 'Sb', 'Te', 'I', 'Xe', \
					'Cs', 'Ba', 'La', 'Ce', 'Pr', 'Nd', 'Pm', 'Sm', 'Eu', 'Gd', 'Tb', 'Dy', 'Ho', 'Er', 'Ym', 'Yb', 'Lu', 'Ha', 'Ta', \
					'W', 'Re', 'Os', 'Ir', 'Pt', 'Au', 'Hg', 'Tl', 'Pb', 'Bi', 'Po', 'At', 'Rn', 'Fr', 'Ra', 'Ac', 'Th', 'Pa', 'U', \
					'Np', 'Pu', 'Am', 'Cm', 'Bk', 'Cf', 'Es', 'Fm', 'Md', 'No', 'Lr', 'Rf', 'Db', 'Sg', 'Bh', 'Hs', 'Mt', 'Ds', 'Rg', \
					'Cn', 'Nh', 'Fl', 'Mc', 'Lv', 'Ts', 'Og']
	element = periodTable[eleNumber]
	return element

# ========================== Read original output file ==========================
print("Please specify the Gaussian output file path:")

# For Unix/Linux OS
fileName = input("(e.g.: /CSIgen/example/DR3b_CS2021.log)\n")
if fileName.strip()[0] == '\'' and fileName.strip()[-1] == '\'':
    fileName = fileName.strip()[1:-1]

# For Microsoft Windows
#fileName = input("(e.g.: C:\\CSIgen\\example\\DR3b_CS2021.log)\n")
#if fileName.strip()[0] == '\"' and fileName.strip()[-1] == '\"':
#    fileName = fileName.strip()[1:-1]

print("\nProcessing output file, please wait...\n")

slashNo = 0
title1 = fileName.strip()
for c in range(len(title1)):
	if title1[c] == '\\' or title1[c] == '/':
		slashNo = c

title2 = title1[slashNo + 1:len(title1) - 4]

with open(fileName.strip(), 'r') as output:
	outputFile = output.readlines()

for i in range(len(outputFile)):
	if '------' in outputFile[i] and '#' in outputFile[i + 1]:
		routeLine = outputFile[i + 1].strip().lower()
		if '------' not in outputFile[i + 2]:
			routeLine = outputFile[i + 1].strip().lower() + ' ' + outputFile[i + 2].strip().lower()
		if 'opt' in routeLine and 'freq' in routeLine:
			jobType = 'opt+freq'
		elif 'opt' in routeLine:
			jobType = 'opt'
		elif 'freq' in routeLine:
			jobType = 'freq'
		else:
			jobType = 'other'
		break

geoFlagStart = 0

for j in range(len(outputFile)):
	if 'Charge =' in outputFile[j] and 'Multiplicity =' in outputFile[j]:
		charge = int(outputFile[j].split()[2])
		multiplicity = int(outputFile[j].split()[5])
	if 'Full point group' in outputFile[j]:
		pointGroup = outputFile[j].split()[3]
	if 'Standard orientation:' in outputFile[j]:
		geoFlagStart = j
if geoFlagStart == 0:
	for k in range(len(outputFile)):
		if 'Input orientation:' in outputFile[k]:
			geoFlagStart = k

for m in range(geoFlagStart + 5, len(outputFile)):
	if '------' in outputFile[m]:
		geoFlagEnd = m
		break

coorLines = []                   # all of the coordinates line are saved in this list, no modified, direct extracting from output

for coorLine in outputFile[geoFlagStart + 5 : geoFlagEnd]:
	coorLines.append(coorLine.strip())

for l in range(len(outputFile)):
	if 'SCF Done' in outputFile[l]:
		eleEnergy = format(float(outputFile[l].split()[4]), '.6f')
	if 'zero-point Energies=' in outputFile[l]:
		freqZPE = format(float(outputFile[l].split()[6]), '.6f')
		freqThr = format(float(outputFile[l + 1].split()[6]), '.6f')
		freqH = format(float(outputFile[l + 2].split()[6]), '.6f')
		freqFE = format(float(outputFile[l + 3].split()[7]), '.6f')

if jobType == 'opt+freq' or jobType == 'freq':
	freqValue = []
	for o in range(len(outputFile)):
		if 'Harmonic frequencies' in outputFile[o]:
			freqStartLine = o
	for p in range(freqStartLine + 6, len(outputFile)):
		if 'Frequencies --' in outputFile[p]:
			freqValue.append(format(float(outputFile[p].split()[2]), '.2f'))
			freqValue.append(format(float(outputFile[p].split()[3]), '.2f'))
			freqValue.append(format(float(outputFile[p].split()[4]), '.2f'))
	imFreq = 0
	for q in range(len(freqValue)):
		if float(freqValue[q]) < 0.0:
			imFreq += 1

print("             SI Style")
print("----------------------------------")
print("   1 - Full (.txt)")
print("   2 - Simple (.txt)")
print("   3 - Coordinates only (.txt)")
print("   4 - Full (.xlsx)")
print("   5 - Simple (.xlsx)")
print("   6 - Coordinates only (.xlsx)")
print("----------------------------------")
csiType = int(input("Please input style number: "))
print("\nWritting SI data, please wait...\n")

normalTer = 1

if csiType == 1 or csiType == 2 or csiType == 3:
	csiOut = open(f'{fileName.strip()[:-3]}txt', 'w')
	csiOut.write(f'{title2}\n')
	if csiType == 1 or csiType == 2:
		csiOut.write(f'\n{routeLine}\n')
		csiOut.write(f'Charge = {charge}, Multiplicity = {multiplicity}, Point group = {pointGroup}\n')
		csiOut.write(f'Electronic Energy = {eleEnergy}\n')
	if csiType == 1:
		if jobType == 'opt+freq' or jobType == 'freq':
			csiOut.write(f'Number of imaginary frequencies = {imFreq}')
			if imFreq == 0:
				csiOut.write('\n')
			else:
				csiOut.write(f', vi = {freqValue[0]}\n')
			csiOut.write(f'Sum of electronic and zero-point Energies = {freqZPE}\n')
			csiOut.write(f'Sum of electronic and thermal Energies = {freqThr}\n')
			csiOut.write(f'Sum of electronic and thermal Enthalpies = {freqH}\n')
			csiOut.write(f'Sum of electronic and thermal Free Energies = {freqFE}\n')
	csiOut.write('\n---------------------------------------------------\n')
	csiOut.write('                  Coordinates (Angstroms)\n')
	csiOut.write(' Atoms        X              Y              Z\n')
	csiOut.write('---------------------------------------------------\n')
	for n in range(len(coorLines)):
		csiOut.write(f'   {elementNo(int(coorLines[n].split()[1]))}')
		if coorLines[n].split()[3][0] == '-':
			csiOut.write(f'      {coorLines[n].split()[3]}')
		else:
			csiOut.write(f'       {coorLines[n].split()[3]}')
		if coorLines[n].split()[4][0] == '-':
			csiOut.write(f'      {coorLines[n].split()[4]}')
		else:
			csiOut.write(f'       {coorLines[n].split()[4]}')
		if coorLines[n].split()[5][0] == '-':
			csiOut.write(f'      {coorLines[n].split()[5]}\n')
		else:
			csiOut.write(f'       {coorLines[n].split()[5]}\n')
	csiOut.write('---------------------------------------------------\n')
	csiOut.close()

# Excel output section
elif csiType == 4:
	outWB = openpyxl.Workbook()
	outWS = outWB.active
	mediumB = Side(border_style = 'medium', color = '000000')
	thinB = Side(border_style = 'thin', color = '000000')
	dashB = Side(border_style = 'dashed', color = '000000')

	outWS.column_dimensions['A'].width = 7.25
	outWS.column_dimensions['B'].width = 11.25
	outWS.column_dimensions['C'].width = 11.25
	outWS.column_dimensions['D'].width = 11.25
	outWS.column_dimensions['E'].width = 7.25
	outWS.column_dimensions['F'].width = 11.25
	outWS.column_dimensions['G'].width = 11.25
	outWS.column_dimensions['H'].width = 11.25
	for w in range(1, 5000):
		outWS.row_dimensions[w].height = 16.0

	cellRange3 = outWS['A1:H1']
	for cellNum3 in cellRange3:
		for cellNum4 in cellNum3:
			cellNum4.border = Border(bottom = mediumB)

	cellRange5 = outWS['A9:H9']
	for cellNum5 in cellRange5:
		for cellNum6 in cellNum5:
			cellNum6.border = Border(bottom = thinB)

	cellRange = outWS['A2:H9']
	for cellNum in cellRange:
		for cellNum1 in cellNum:
			cellNum1.font = Font(name = 'Times New Roman', size = 10)

	outWS.merge_cells('A1:H1')
	outWS['A1'] = title2
	outWS['A1'].font = Font(name = 'Times New Roman', size = 10.5, bold = True)
	outWS['A1'].alignment = Alignment(horizontal = 'center', vertical = 'center')

	outWS.merge_cells('A2:C9')
	outWS['A2'].font = Font(name = 'Times New Roman', size = 10.5, color = '929292')
	outWS['A2'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['A2'] = 'Insert molecular geometry here.'

	outWS.merge_cells('D2:H2')
	outWS['D2'] = routeLine
	outWS['D2'].font = Font(name = 'Courier', size = 10)

	outWS.merge_cells('D3:H3')
	outWS['D3'] = f'Charge = {charge}, Multiplicity = {multiplicity}, Point group = {pointGroup}'

	outWS.merge_cells('D4:H4')
	outWS['D4'] = f'Electronic Energy = {eleEnergy} Hartree'

	outWS.merge_cells('D5:H5')
	if jobType == 'opt+freq' or jobType == 'freq':
		if imFreq == 0:
			outWS['D5'] = f'Number of imaginary frequencies = 0'
		else:
			outWS['D5'] = f'Number of imaginary frequencies = {imFreq}, vi = {freqValue[0]}'

	outWS.merge_cells('D6:H6')
	outWS.merge_cells('D7:H7')
	outWS.merge_cells('D8:H8')
	outWS.merge_cells('D9:H9')
	if jobType == 'opt+freq' or jobType == 'freq':
		outWS['D6'] = f'Sum of electronic and zero-point Energies = {freqZPE} Hartree'
		outWS['D7'] = f'Sum of electronic and thermal Energies = {freqThr} Hartree'
		outWS['D8'] = f'Sum of electronic and thermal Enthalpies = {freqH} Hartree'
		outWS['D9'] = f'Sum of electronic and thermal Free Energies = {freqFE} Hartree'

	outWS.merge_cells('A10:A11')
	outWS.merge_cells('B10:D10')
	outWS.merge_cells('E10:E11')
	outWS.merge_cells('F10:H10')

	outWS['A10'] = 'Atoms'
	outWS['A10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
	outWS['A10'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
	outWS['A11'].border = Border(bottom = thinB)

	outWS['B10'] = 'Cartesian Coordinates'
	outWS['B10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
	outWS['B10'].alignment = Alignment(horizontal = 'center', vertical = 'center')

	outWS['B11'] = 'X'
	outWS['B11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['B11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['B11'].border = Border(top = dashB, bottom = thinB)
	outWS['C11'] = 'Y'
	outWS['C11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['C11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['C11'].border = Border(top = dashB, bottom = thinB)
	outWS['D11'] = 'Z'
	outWS['D11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['D11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['D11'].border = Border(top = dashB, bottom = thinB)

	outWS['E10'] = 'Atoms'
	outWS['E10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
	outWS['E10'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
	outWS['E11'].border = Border(bottom = thinB)

	outWS['F10'] = 'Cartesian Coordinates'
	outWS['F10'].font = Font(name = 'Times New Roman', size = 10, bold = True)
	outWS['F10'].alignment = Alignment(horizontal = 'center', vertical = 'center')

	outWS['F11'] = 'X'
	outWS['F11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['F11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['F11'].border = Border(top = dashB, bottom = thinB)
	outWS['G11'] = 'Y'
	outWS['G11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['G11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['G11'].border = Border(top = dashB, bottom = thinB)
	outWS['H11'] = 'Z'
	outWS['H11'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['H11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['H11'].border = Border(top = dashB, bottom = thinB)

	excelCoors = []
	if len(coorLines) % 2 == 0:
		excelCoorLineNO = int(len(coorLines) / 2)
		for r in range(0, len(coorLines), 2):
			excelCoor = []
			excelCoor.append(elementNo(int(coorLines[r].split()[1])))
			excelCoor.append(coorLines[r].split()[3])
			excelCoor.append(coorLines[r].split()[4])
			excelCoor.append(coorLines[r].split()[5])
			excelCoor.append(elementNo(int(coorLines[r + 1].split()[1])))
			excelCoor.append(coorLines[r + 1].split()[3])
			excelCoor.append(coorLines[r + 1].split()[4])
			excelCoor.append(coorLines[r + 1].split()[5])
			excelCoors.append(excelCoor)

	elif len(coorLines) % 2 == 1:
		excelCoorLineNO = int((len(coorLines) + 1) / 2)
		for r in range(0, len(coorLines) - 1, 2):
			excelCoor = []
			excelCoor.append(elementNo(int(coorLines[r].split()[1])))
			excelCoor.append(coorLines[r].split()[3])
			excelCoor.append(coorLines[r].split()[4])
			excelCoor.append(coorLines[r].split()[5])
			excelCoor.append(elementNo(int(coorLines[r + 1].split()[1])))
			excelCoor.append(coorLines[r + 1].split()[3])
			excelCoor.append(coorLines[r + 1].split()[4])
			excelCoor.append(coorLines[r + 1].split()[5])
			excelCoors.append(excelCoor)
		excelCoor = []
		excelCoor.append(elementNo(int(coorLines[-1].split()[1])))
		excelCoor.append(coorLines[-1].split()[3])
		excelCoor.append(coorLines[-1].split()[4])
		excelCoor.append(coorLines[-1].split()[5])
		excelCoors.append(excelCoor)

	for line in excelCoors:
		outWS.append(line)

	cellRange2 = outWS['A12:H4001']
	for cellNum in cellRange2:
		for cellNum1 in cellNum:
			cellNum1.font = Font(name = 'Times New Roman', size = 10)
			cellNum1.alignment = Alignment(horizontal = 'center', vertical = 'center')

	lastLine = excelCoorLineNO + 11

	cellRange7 = outWS[f'A{lastLine}:H{lastLine}']
	for cellNum7 in cellRange7:
		for cellNum8 in cellNum7:
			cellNum8.border = Border(bottom = mediumB)
	cellRange9 = outWS[f'E10:E{lastLine}']
	for cellNum9 in cellRange9:
		for cellNum10 in cellNum9:
			cellNum10.border = Border(left = dashB)
	outWS['E11'].border = Border(left = dashB, bottom = thinB)
	outWS[f'E{lastLine}'].border = Border(left = dashB, bottom = mediumB)

	outWB.save(f'{fileName.strip()[:-3]}xlsx')

elif csiType == 5:
	outWB = openpyxl.Workbook()
	outWS = outWB.active
	mediumB = Side(border_style = 'medium', color = '000000')
	thinB = Side(border_style = 'thin', color = '000000')
	dashB = Side(border_style = 'dashed', color = '000000')

	outWS.column_dimensions['A'].width = 8.25
	outWS.column_dimensions['B'].width = 12.25
	outWS.column_dimensions['C'].width = 12.25
	outWS.column_dimensions['D'].width = 12.25

	cellRange = outWS['A2:D4']
	for cellNum in cellRange:
		for cellNum1 in cellNum:
			cellNum1.font = Font(name = 'Times New Roman', size = 10)

	for w in range(1, 8000):
		outWS.row_dimensions[w].height = 16.0

	cellRange3 = outWS['A1:D1']
	for cellNum3 in cellRange3:
		for cellNum4 in cellNum3:
			cellNum4.border = Border(bottom = mediumB)

	cellRange5 = outWS['A4:D4']
	for cellNum5 in cellRange5:
		for cellNum6 in cellNum5:
			cellNum6.border = Border(bottom = thinB)

	outWS.merge_cells('A1:D1')
	outWS['A1'] = title2
	outWS['A1'].font = Font(name = 'Times New Roman', size = 10.5, bold = True)
	outWS['A1'].alignment = Alignment(horizontal = 'center', vertical = 'center')

	outWS.merge_cells('A2:D2')
	outWS['A2'] = routeLine
	outWS['A2'].font = Font(name = 'Courier', size = 10)

	outWS.merge_cells('A3:D3')
	outWS['A3'] = f'Charge = {charge}, Multiplicity = {multiplicity}'

	outWS.merge_cells('A4:D4')
	outWS['A4'] = f'Electronic Energy = {eleEnergy} Hartree'

	outWS.merge_cells('A5:A6')
	outWS.merge_cells('B5:D5')

	outWS['A5'] = 'Atoms'
	outWS['A5'].font = Font(name = 'Times New Roman', size = 10, bold = True)
	outWS['A5'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
	outWS['A6'].border = Border(bottom = thinB)

	outWS['B5'] = 'Cartesian Coordinates'
	outWS['B5'].font = Font(name = 'Times New Roman', size = 10, bold = True)
	outWS['B5'].alignment = Alignment(horizontal = 'center', vertical = 'center')

	outWS['B6'] = 'X'
	outWS['B6'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['B6'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['B6'].border = Border(top = dashB, bottom = thinB)
	outWS['C6'] = 'Y'
	outWS['C6'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['C6'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['C6'].border = Border(top = dashB, bottom = thinB)
	outWS['D6'] = 'Z'
	outWS['D6'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['D6'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['D6'].border = Border(top = dashB, bottom = thinB)

	excelCoors = []
	for s in range(len(coorLines)):
		excelCoor = []
		excelCoor.append(elementNo(int(coorLines[s].split()[1])))
		excelCoor.append(coorLines[s].split()[3])
		excelCoor.append(coorLines[s].split()[4])
		excelCoor.append(coorLines[s].split()[5])
		excelCoors.append(excelCoor)

	for line in excelCoors:
		outWS.append(line)
	
	cellRange1 = outWS['A7:D8007']
	for cellNum in cellRange1:
		for cellNum1 in cellNum:
			cellNum1.font = Font(name = 'Times New Roman', size = 10)
			cellNum1.alignment = Alignment(horizontal = 'center', vertical = 'center')

	lastLine = len(coorLines) + 6

	cellRange7 = outWS[f'A{lastLine}:D{lastLine}']
	for cellNum7 in cellRange7:
		for cellNum8 in cellNum7:
			cellNum8.border = Border(bottom = mediumB)

	outWB.save(f'{fileName.strip()[:-3]}xlsx')


elif csiType == 6:
	outWB = openpyxl.Workbook()
	outWS = outWB.active
	mediumB = Side(border_style = 'medium', color = '000000')
	thinB = Side(border_style = 'thin', color = '000000')
	dashB = Side(border_style = 'dashed', color = '000000')

	outWS.column_dimensions['A'].width = 7.25
	outWS.column_dimensions['B'].width = 11.25
	outWS.column_dimensions['C'].width = 11.25
	outWS.column_dimensions['D'].width = 11.25

	for w in range(1, 8000):
		outWS.row_dimensions[w].height = 16.0

	cellRange3 = outWS['A1:D1']
	for cellNum3 in cellRange3:
		for cellNum4 in cellNum3:
			cellNum4.border = Border(bottom = mediumB)

	outWS.merge_cells('A1:D1')
	outWS['A1'] = title2
	outWS['A1'].font = Font(name = 'Times New Roman', size = 10.5, bold = True)
	outWS['A1'].alignment = Alignment(horizontal = 'center', vertical = 'center')

	outWS.merge_cells('A2:A3')
	outWS.merge_cells('B2:D2')

	outWS['A2'] = 'Atoms'
	outWS['A2'].font = Font(name = 'Times New Roman', size = 10, bold = True)
	outWS['A2'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
	outWS['A3'].border = Border(bottom = thinB)

	outWS['B2'] = 'Cartesian Coordinates'
	outWS['B2'].font = Font(name = 'Times New Roman', size = 10, bold = True)
	outWS['B2'].alignment = Alignment(horizontal = 'center', vertical = 'center')

	outWS['B3'] = 'X'
	outWS['B3'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['B3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['B3'].border = Border(top = dashB, bottom = thinB)
	outWS['C3'] = 'Y'
	outWS['C3'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['C3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['C3'].border = Border(top = dashB, bottom = thinB)
	outWS['D3'] = 'Z'
	outWS['D3'].font = Font(name = 'Times New Roman', size = 10, bold = True, italic = True)
	outWS['D3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
	outWS['D3'].border = Border(top = dashB, bottom = thinB)

	excelCoors = []
	for s in range(len(coorLines)):
		excelCoor = []
		excelCoor.append(elementNo(int(coorLines[s].split()[1])))
		excelCoor.append(coorLines[s].split()[3])
		excelCoor.append(coorLines[s].split()[4])
		excelCoor.append(coorLines[s].split()[5])
		excelCoors.append(excelCoor)

	for line in excelCoors:
		outWS.append(line)
	
	cellRange1 = outWS['A4:D8005']
	for cellNum in cellRange1:
		for cellNum1 in cellNum:
			cellNum1.font = Font(name = 'Times New Roman', size = 10)
			cellNum1.alignment = Alignment(horizontal = 'center', vertical = 'center')

	lastLine = len(coorLines) + 3

	cellRange7 = outWS[f'A{lastLine}:D{lastLine}']
	for cellNum7 in cellRange7:
		for cellNum8 in cellNum7:
			cellNum8.border = Border(bottom = mediumB)

	outWB.save(f'{fileName.strip()[:-3]}xlsx')

else:
	print("Input error, CSIgen would be terminated.\n")
	normalTer = 0

# ========================== Result information ==========================
print("*******************************************************************************")
print("")
if normalTer == 1:
	print("                    SI file has been saved at same dictionary.")
else:
	print("                  SI file was not saved due to error inputting.")
print("                         Normal termination of CSIgen.")
print("")
print("*******************************************************************************\n")

